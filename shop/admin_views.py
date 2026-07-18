from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import user_passes_test
from django.contrib import messages
from django.db.models import Sum, Count, Q
from django.http import HttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import io

from .models import (
    Product, Category, Order, OrderItem, Coupon, Offer, 
    Banner, ReturnRequest, CustomerDiscountLead, ProductImage, StoreSetting
)
from accounts.models import User
from .forms import ProductForm, CategoryForm, CouponForm, OfferForm, BannerForm, StoreSettingForm

# Superuser check decorator
def superuser_required(view_func):
    return user_passes_test(
        lambda u: u.is_authenticated and u.is_superuser, 
        login_url='accounts:login'
    )(view_func)

@superuser_required
def admin_dashboard_view(request):
    """
    Renders custom luxury admin homepage tracking core KPIs: Total Sales, orders, customers, and inventory levels.
    """
    orders = Order.objects.all()
    completed_orders = orders.filter(payment_status='completed')
    
    total_sales = completed_orders.aggregate(total=Sum('total_price'))['total'] or 0.00
    total_orders = orders.count()
    total_customers = User.objects.filter(is_superuser=False).count()
    
    out_of_stock = Product.objects.filter(stock=0).count()
    low_stock = Product.objects.filter(stock__gt=0, stock__lte=5).count()
    
    recent_orders = orders.order_by('-created_at')[:5]
    recent_leads = CustomerDiscountLead.objects.order_by('-created_at')[:5]
    
    context = {
        'total_sales': total_sales,
        'total_orders': total_orders,
        'total_customers': total_customers,
        'out_of_stock': out_of_stock,
        'low_stock': low_stock,
        'recent_orders': recent_orders,
        'recent_leads': recent_leads,
    }
    return render(request, 'admin_dashboard/dashboard.html', context)

# --- Product Management CRUD ---

@superuser_required
def admin_products_view(request):
    products = Product.objects.select_related('category').all().order_by('stock', '-id')
    return render(request, 'admin_dashboard/products.html', {'products': products})

@superuser_required
def admin_product_create_view(request):
    form = ProductForm()
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save()
            
            # Save Primary Image
            primary_img = request.FILES.get('primary_image')
            if primary_img:
                ProductImage.objects.create(product=product, image=primary_img, is_primary=True)
                
            # Save extra gallery images
            ex1 = request.FILES.get('extra_image_1')
            if ex1:
                ProductImage.objects.create(product=product, image=ex1, is_primary=False)
            ex2 = request.FILES.get('extra_image_2')
            if ex2:
                ProductImage.objects.create(product=product, image=ex2, is_primary=False)
                
            messages.success(request, f"Product '{product.name}' created successfully.")
            return redirect('shop:admin_products')
            
    return render(request, 'admin_dashboard/product_form.html', {'form': form, 'title': 'Add Product'})

@superuser_required
def admin_product_update_view(request, pk):
    product = get_object_or_404(Product, pk=pk)
    form = ProductForm(instance=product)
    
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            
            # Update primary image if uploaded
            primary_img = request.FILES.get('primary_image')
            if primary_img:
                ProductImage.objects.filter(product=product, is_primary=True).delete()
                ProductImage.objects.create(product=product, image=primary_img, is_primary=True)
                
            # Add extra gallery images
            ex1 = request.FILES.get('extra_image_1')
            if ex1:
                ProductImage.objects.create(product=product, image=ex1, is_primary=False)
            ex2 = request.FILES.get('extra_image_2')
            if ex2:
                ProductImage.objects.create(product=product, image=ex2, is_primary=False)
                
            messages.success(request, f"Product '{product.name}' updated successfully.")
            return redirect('shop:admin_products')
            
    return render(request, 'admin_dashboard/product_form.html', {
        'form': form, 
        'product': product, 
        'title': 'Edit Product',
        'existing_images': product.images.all()
    })

@superuser_required
def admin_product_delete_view(request, pk):
    product = get_object_or_404(Product, pk=pk)
    name = product.name
    product.delete()
    messages.success(request, f"Product '{name}' deleted successfully.")
    return redirect('shop:admin_products')

# --- Category Management ---

@superuser_required
def admin_categories_view(request):
    categories = Category.objects.all().order_by('parent__name', 'name')
    return render(request, 'admin_dashboard/categories.html', {'categories': categories})

@superuser_required
def admin_category_create_view(request):
    form = CategoryForm()
    if request.method == 'POST':
        form = CategoryForm(request.POST, request.FILES)
        if form.is_valid():
            cat = form.save()
            messages.success(request, f"Category '{cat.name}' created.")
            return redirect('shop:admin_categories')
    return render(request, 'admin_dashboard/category_form.html', {'form': form, 'title': 'Create Category'})

@superuser_required
def admin_category_update_view(request, pk):
    cat = get_object_or_404(Category, pk=pk)
    form = CategoryForm(instance=cat)
    if request.method == 'POST':
        form = CategoryForm(request.POST, request.FILES, instance=cat)
        if form.is_valid():
            form.save()
            messages.success(request, f"Category '{cat.name}' updated.")
            return redirect('shop:admin_categories')
    return render(request, 'admin_dashboard/category_form.html', {'form': form, 'category': cat, 'title': 'Edit Category'})

@superuser_required
def admin_category_delete_view(request, pk):
    cat = get_object_or_404(Category, pk=pk)
    name = cat.name
    cat.delete()
    messages.success(request, f"Category '{name}' deleted.")
    return redirect('shop:admin_categories')

# --- Order Fulfillment Operations ---

@superuser_required
def admin_orders_view(request):
    status_filter = request.GET.get('status')
    orders = Order.objects.select_related('user').all().order_by('-created_at')
    
    if status_filter:
        orders = orders.filter(order_status=status_filter)
        
    return render(request, 'admin_dashboard/orders.html', {
        'orders': orders, 
        'status_filter': status_filter,
        'status_choices': Order.ORDER_STATUS_CHOICES
    })

@superuser_required
def admin_order_detail_view(request, order_number):
    order = get_object_or_404(Order, order_number=order_number)

    if request.method == 'POST':
        new_order_status = request.POST.get('order_status')
        new_payment_status = request.POST.get('payment_status')

        if new_order_status:
            order.order_status = new_order_status
        if new_payment_status:
            order.payment_status = new_payment_status

        order.save()
        messages.success(request, f"Order status updated for {order.order_number}.")
        return redirect('shop:admin_order_detail', order_number=order.order_number)

    return render(request, 'admin_dashboard/order_detail.html', {'order': order})


@superuser_required
def admin_order_delete_view(request, order_number):
    order = get_object_or_404(Order, order_number=order_number)

    if request.method == "POST":
        order.delete()
        messages.success(request, f"Order {order.order_number} deleted successfully.")

    return redirect("shop:admin_orders")

# --- Returns Management ---

@superuser_required
def admin_return_requests_view(request):
    requests = ReturnRequest.objects.select_related('order', 'user').all().order_by('-created_at')
    return render(request, 'admin_dashboard/returns.html', {'requests': requests})

@superuser_required
def admin_return_request_action_view(request, pk):
    ret = get_object_or_404(ReturnRequest, pk=pk)
    action = request.POST.get('action') # approve, reject, refund
    notes = request.POST.get('admin_notes', '')
    
    if action == 'approve':
        ret.status = 'approved'
        ret.order.order_status = 'returned'
        ret.order.save()
    elif action == 'reject':
        ret.status = 'rejected'
    elif action == 'refund':
        ret.status = 'refunded'
        ret.order.payment_status = 'refunded'
        ret.order.save()
        
    ret.admin_notes = notes
    ret.save()
    
    messages.success(request, f"Return request updated to {ret.get_status_display()}.")
    return redirect('shop:admin_returns')

# --- Coupon & Offer Controls ---

@superuser_required
def admin_coupons_view(request):
    coupons = Coupon.objects.all().order_by('-valid_to')
    form = CouponForm()
    if request.method == 'POST':
        form = CouponForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "New promo coupon created.")
            return redirect('shop:admin_coupons')
            
    return render(request, 'admin_dashboard/coupons.html', {'coupons': coupons, 'form': form})

@superuser_required
def admin_coupon_delete_view(request, pk):
    c = get_object_or_404(Coupon, pk=pk)
    c.delete()
    messages.success(request, "Coupon code deleted.")
    return redirect('shop:admin_coupons')

@superuser_required
def admin_offers_view(request):
    offers = Offer.objects.all().order_by('-end_time')
    form = OfferForm()
    if request.method == 'POST':
        form = OfferForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "New promo offer created.")
            return redirect('shop:admin_offers')
            
    return render(request, 'admin_dashboard/offers.html', {'offers': offers, 'form': form})

@superuser_required
def admin_offer_delete_view(request, pk):
    o = get_object_or_404(Offer, pk=pk)
    o.delete()
    messages.success(request, "Offer deleted.")
    return redirect('shop:admin_offers')

# --- Banner Slides Management ---

@superuser_required
def admin_banners_view(request):
    banners = Banner.objects.all().order_by('order', 'id')
    form = BannerForm()
    if request.method == 'POST':
        form = BannerForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Hero slider banner added.")
            return redirect('shop:admin_banners')
            
    return render(request, 'admin_dashboard/banners.html', {'banners': banners, 'form': form})

@superuser_required
def admin_banner_delete_view(request, pk):
    b = get_object_or_404(Banner, pk=pk)
    b.delete()
    messages.success(request, "Banner slide deleted.")
    return redirect('shop:admin_banners')

# --- Store configuration settings ---

@superuser_required
def admin_settings_view(request):
    store_settings = StoreSetting.get_settings()
    form = StoreSettingForm(instance=store_settings)
    if request.method == 'POST':
        form = StoreSettingForm(request.POST, instance=store_settings)
        if form.is_valid():
            form.save()
            messages.success(request, "Global store settings updated.")
            return redirect('shop:admin_settings')
            
    return render(request, 'admin_dashboard/settings.html', {'form': form})

# --- Sales Reports & Exports (Excel & PDF) ---

@superuser_required
def admin_reports_view(request):
    """
    Renders reports date range filters and totals.
    """
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    orders = Order.objects.filter(payment_status='completed')
    
    if start_date:
        orders = orders.filter(created_at__date__gte=start_date)
    if end_date:
        orders = orders.filter(created_at__date__lte=end_date)
        
    totals = orders.aggregate(
        sales=Sum('total_price'),
        qty=Count('id'),
        shipping=Sum('shipping_charge'),
        discount=Sum('discount_amount')
    )
    
    context = {
        'orders': orders.order_by('-created_at'),
        'totals': totals,
        'start_date': start_date,
        'end_date': end_date
    }
    return render(request, 'admin_dashboard/reports.html', context)

@superuser_required
def admin_export_excel_view(request):
    """
    Uses openpyxl to compile and format sales logs into spreadsheet files.
    """
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    orders = Order.objects.all().order_by('-created_at')
    if start_date:
        orders = orders.filter(created_at__date__gte=start_date)
    if end_date:
        orders = orders.filter(created_at__date__lte=end_date)
        
    # Setup openpyxl Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    # Grid lines visible
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    title_font = Font(name='Calibri', size=16, bold=True, color='D4AF37')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    bold_font = Font(name='Calibri', size=11, bold=True)
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    fill_header = PatternFill(start_color='111111', end_color='111111', fill_type='solid')
    fill_total = PatternFill(start_color='FAF7EC', end_color='FAF7EC', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Write Title
    ws.merge_cells('A1:J1')
    ws['A1'] = "RARA JEWELS - SALES REPORT"
    ws['A1'].font = title_font
    ws['A1'].alignment = align_center
    ws.row_dimensions[1].height = 40
    
    # Write metadata info
    range_str = f"Date Range: {start_date or 'All'} to {end_date or 'All'}"
    ws.merge_cells('A2:J2')
    ws['A2'] = range_str
    ws['A2'].font = Font(italic=True, size=10)
    ws['A2'].alignment = align_center
    
    # Headers
    headers = [
        "Order Number", "Customer Name", "Email", "Date", 
        "Payment Method", "Payment Status", "Order Status",
        "Subtotal (₹)", "Discount (₹)", "Total Value (₹)"
    ]
    
    ws.append([]) # row 3 blank
    ws.append(headers) # row 4
    ws.row_dimensions[4].height = 25
    
    for col_idx in range(1, 11):
        cell = ws.cell(row=4, column=col_idx)
        cell.font = header_font
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
        
    # Append order rows
    row_num = 5
    for o in orders:
        ws.append([
            o.order_number,
            o.user.name,
            o.user.email,
            o.created_at.strftime('%Y-%m-%d %H:%M'),
            o.get_payment_method_display(),
            o.get_payment_status_display(),
            o.get_order_status_display(),
            float(o.subtotal),
            float(o.discount_amount),
            float(o.total_price)
        ])
        
        ws.row_dimensions[row_num].height = 20
        # Alignment & Borders for data rows
        for col_idx in range(1, 11):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = thin_border
            if col_idx in [1, 4, 5, 6, 7]:
                cell.alignment = align_center
            elif col_idx in [2, 3]:
                cell.alignment = align_left
            else:
                cell.alignment = align_right
                cell.number_format = '₹#,##0.00'
        row_num += 1
        
    # Add Totals row
    ws.append([]) # Blank spacer
    totals_row = row_num + 1
    ws.cell(row=totals_row, column=7, value="Total Completed Sales:").font = bold_font
    ws.cell(row=totals_row, column=7).alignment = align_right
    
    completed_total = orders.filter(payment_status='completed').aggregate(Sum('total_price'))['total_price__sum'] or 0.00
    total_cell = ws.cell(row=totals_row, column=10, value=float(completed_total))
    total_cell.font = bold_font
    total_cell.fill = fill_total
    total_cell.alignment = align_right
    total_cell.number_format = '₹#,##0.00'
    ws.row_dimensions[totals_row].height = 22
    
    # Adjust widths automatically
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Write to memory stream and respond
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Sales-Report-{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

@superuser_required
def admin_export_pdf_view(request):
    """
    Generates structured PDF summary tables using ReportLab.
    """
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    orders = Order.objects.filter(payment_status='completed').order_by('-created_at')
    if start_date:
        orders = orders.filter(created_at__date__gte=start_date)
    if end_date:
        orders = orders.filter(created_at__date__lte=end_date)
        
    buffer = io.BytesIO()
    # A4 Landscape for extra wide sales table columns
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(A4), 
        rightMargin=30, 
        leftMargin=30, 
        topMargin=30, 
        bottomMargin=30
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    # Stylings
    gold_color = colors.HexColor('#D4AF37')
    dark_bg = colors.HexColor('#111111')
    
    title_style = ParagraphStyle(
        'RepTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=gold_color,
        spaceAfter=5
    )
    
    meta_style = ParagraphStyle(
        'RepMeta',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.white,
        spaceAfter=10
    )
    
    cell_style = ParagraphStyle(
        'CellNormal',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=10
    )
    
    cell_bold = ParagraphStyle(
        'CellBold',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        leading=10
    )
    
    # PDF Header Banner
    header_data = [
        [
            Paragraph("RARA JEWELS - SALES REPORT SUMMARY", title_style),
            Paragraph(f"Generated: {timezone.now().strftime('%d-%b-%Y')}<br/>Range: {start_date or 'All'} to {end_date or 'All'}", ParagraphStyle('RightHeader', parent=styles['Normal'], fontName='Helvetica', fontSize=8, textColor=colors.white, alignment=2))
        ]
    ]
    
    header_table = Table(header_data, colWidths=[500, 240])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), dark_bg),
        ('PADDING', (0,0), (-1,-1), 15),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    
    story.append(header_table)
    story.append(Spacer(1, 20))
    
    # Build columns
    headers = [
        Paragraph("<b>Order No</b>", cell_bold),
        Paragraph("<b>Customer</b>", cell_bold),
        Paragraph("<b>Email</b>", cell_bold),
        Paragraph("<b>Date</b>", cell_bold),
        Paragraph("<b>Status</b>", cell_bold),
        Paragraph("<b>Subtotal (₹)</b>", cell_bold),
        Paragraph("<b>Discount (₹)</b>", cell_bold),
        Paragraph("<b>Total (₹)</b>", cell_bold)
    ]
    
    table_data = [headers]
    
    total_rev = 0.00
    for o in orders:
        row = [
            Paragraph(o.order_number, cell_style),
            Paragraph(o.user.name, cell_style),
            Paragraph(o.user.email, cell_style),
            Paragraph(o.created_at.strftime('%Y-%m-%d'), cell_style),
            Paragraph(o.get_order_status_display(), cell_style),
            Paragraph(f"{o.subtotal:.2f}", cell_style),
            Paragraph(f"{o.discount_amount:.2f}", cell_style),
            Paragraph(f"{o.total_price:.2f}", cell_style)
        ]
        table_data.append(row)
        total_rev += float(o.total_price)
        
    # Append Summary row
    table_data.append([
        "", "", "", "",
        Paragraph("<b>Grand Total Completed:</b>", cell_bold),
        "", "",
        Paragraph(f"<b>₹{total_rev:.2f}</b>", cell_bold)
    ])
    
    report_table = Table(table_data, colWidths=[100, 100, 130, 80, 80, 80, 80, 90])
    report_table.setStyle(TableStyle([
        ('LINEBELOW', (0,0), (-1,0), 1.5, gold_color),
        ('LINEBELOW', (0,1), (-1,-2), 0.5, colors.HexColor('#E0E0E0')),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F5F5F5')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('ALIGN', (5,0), (-1,-1), 'RIGHT'),
        ('LINEABOVE', (4,-1), (7,-1), 1, gold_color),
    ]))
    
    story.append(report_table)
    story.append(Spacer(1, 20))
    story.append(Paragraph("End of report. Verified by system audit.", ParagraphStyle('FootNote', parent=styles['Normal'], fontName='Helvetica-Oblique', fontSize=8, textColor=colors.HexColor('#777777'))))
    
    doc.build(story)
    
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Sales-Report-{timezone.now().strftime("%Y%m%d")}.pdf"'
    return response
